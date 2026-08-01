#!/usr/bin/env python3
import argparse
import json
import math
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl\n" + str(exc))


SEASONS = ["fog", "shimmer", "storm", "soil", "threshold"]
DEFAULT_SPIRIT = {
    "fog": "The Deer",
    "shimmer": "The Whale",
    "storm": "The Wolf",
    "soil": "The Bear",
    "threshold": "The Witness",
}
DEFAULT_LEGACY = {
    "fog": "The Whisperer",
    "shimmer": "The Lumen Keeper",
    "storm": "The Architect of Ruin",
    "soil": "The Archivist",
    "threshold": "The Threshold Keeper",
}


def normalize_name(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_edges(raw):
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    parts = re.split(r"[,;|/\\]+", text)
    out = []
    for part in parts:
        token = part.strip()
        if token and "openpyxl." not in token and not token.startswith("="):
            out.append(token)
    return out


def build_workbook_index(root):
    idx = {}
    for path in root.rglob("*.xls*"):
        idx[path.name.lower()] = path
    return idx


def open_workbook(path, cache):
    key = str(path)
    if key in cache:
        return cache[key]
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    cache[key] = wb
    return wb


def clean_label(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if text.startswith("="):
        return ""
    if "openpyxl." in text:
        return ""
    return text


def resolve_workbook_path(requested_name, workbook_index, source_priority, root):
    if requested_name:
        direct = workbook_index.get(requested_name.lower())
        if direct:
            return direct

    for preferred in source_priority:
        candidate = workbook_index.get(str(preferred).lower())
        if candidate:
            return candidate

    found = sorted(root.rglob("*.xlsx")) + sorted(root.rglob("*.xlsm"))
    return found[0] if found else None


def sheet_rows(wb, sheet_name, wanted_columns):
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    header_map = {}
    for idx, label in enumerate(header_row):
        key = normalize_name(label)
        if key:
            header_map[key] = idx

    wanted = [c for c in wanted_columns if isinstance(c, str) and c.strip()]
    selected = []
    for col in wanted:
        k = normalize_name(col)
        if k in header_map:
            selected.append((col, header_map[k]))

    if not selected:
        # Fallback: return first eight columns with generated keys.
        selected = [(f"col{i+1}", i) for i in range(min(len(header_row), 8))]

    output = []
    for row in rows:
        rec = {}
        non_empty = False
        for out_key, col_idx in selected:
            value = row[col_idx] if col_idx < len(row) else None
            rec[out_key] = value
            if value not in (None, ""):
                non_empty = True
        if non_empty:
            output.append(rec)

    return output


def classify_season(text):
    lower = str(text or "").lower()
    if any(t in lower for t in ["collapse", "storm", "decay", "rupture", "pressure"]):
        return "storm"
    if any(t in lower for t in ["stable", "anchor", "soil", "root", "calm"]):
        return "soil"
    if any(t in lower for t in ["expand", "shimmer", "bloom", "light", "growth"]):
        return "shimmer"
    if any(t in lower for t in ["drift", "mist", "fog", "transit", "uncertain"]):
        return "fog"
    return "threshold"


def aggregate_numeric(rows, key):
    vals = []
    for row in rows:
        value = parse_float(row.get(key))
        if value is not None:
            vals.append(value)
    if not vals:
        return None
    return sum(vals) / len(vals)


def build_dialogues(data_blocks, source_ref):
    dials = data_blocks.get("Book 1.xlsx::Dials", [])
    engine = data_blocks.get("Stability Engine v0.xlsx::ENGINE", [])

    momentum = aggregate_numeric(dials, "W_Momentum")
    accel = aggregate_numeric(dials, "W_Acceleration")
    slope = aggregate_numeric(dials, "W_Bias_Slope")
    pressure = aggregate_numeric(engine, "PressureScore")

    pressure = pressure if pressure is not None else 0.0
    momentum = momentum if momentum is not None else 0.0
    accel = accel if accel is not None else 0.0
    slope = slope if slope is not None else 0.0

    season = {
        "fog": "The mist remembers beginnings.",
        "shimmer": "Light expands where attention rests.",
        "storm": "Pressure reveals the shape of truth.",
        "soil": "Roots hold what the surface forgets.",
        "threshold": "The threshold listens before it speaks.",
    }

    legacy = {
        "fog": "Whisperer: 'Clarity is a visitor, not a resident.'",
        "shimmer": "Lumen Keeper: 'Expansion without ego becomes illumination.'",
        "storm": "Architect of Ruin: 'Collapse is the doorway to form.'",
        "soil": "Archivist: 'Memory is the slowest kind of light.'",
        "threshold": "Keeper: 'Return is how meaning deepens.'",
    }

    root_machine = {
        "fog": "Drift: 'Movement without direction is still movement.'",
        "shimmer": "Resonance: 'Everything vibrates toward alignment.'",
        "storm": "Collapse: 'Pressure is the architect of change.'",
        "soil": "Anchor: 'Stillness is a kind of strength.'",
        "threshold": "Threshold: 'Presence is the first engine.'",
    }

    voice = {
        "fog": f"Voice channel active: fog (momentum={momentum:.2f}).",
        "shimmer": f"Voice channel active: shimmer (biasSlope={slope:.2f}).",
        "storm": f"Voice channel active: storm (pressure={pressure:.2f}).",
        "soil": f"Voice channel active: soil (accel={accel:.2f}).",
        "threshold": "Voice channel active: threshold.",
    }

    return {
        "season": season,
        "legacy": legacy,
        "rootMachine": root_machine,
        "voice": voice,
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceWorkbook": source_ref,
            "rowCount": len(dials) + len(engine),
        },
    }


def pick_label_by_season(rows, role_keys, defaults):
    buckets = {s: Counter() for s in SEASONS}
    for row in rows:
        seed = " ".join(str(row.get(k) or "") for k in role_keys)
        season = classify_season(seed)
        label = None
        for k in role_keys:
            candidate = clean_label(row.get(k))
            if candidate:
                label = candidate
                break
        if label:
            buckets[season][label] += 1

    output = {}
    for season in SEASONS:
        if buckets[season]:
            output[season] = buckets[season].most_common(1)[0][0]
        else:
            output[season] = defaults[season]
    return output


def build_mythic(data_blocks, source_ref):
    interactions = data_blocks.get("Stability Engine v0.xlsx::INTERACTIONS", [])
    physics = data_blocks.get("Stability Engine v0.xlsx::PHYSICS", [])

    spirit_rows = []
    for row in physics:
        spirit_rows.append({
            "Role": row.get("Role"),
            "State": row.get("State"),
            "Energy": row.get("Energy"),
            "Stability": row.get("Stability"),
        })

    legacy_rows = []
    for row in interactions:
        legacy_rows.append({
            "RoleA": row.get("RoleA"),
            "RoleB": row.get("RoleB"),
            "ResultState": row.get("ResultState"),
            "InteractionType": row.get("InteractionType"),
            "StabilityImpact": row.get("StabilityImpact"),
        })

    spirit = pick_label_by_season(spirit_rows, ["Role", "State"], DEFAULT_SPIRIT)
    legacy = pick_label_by_season(legacy_rows, ["RoleA", "RoleB", "ResultState"], DEFAULT_LEGACY)

    return {
        "spirit": spirit,
        "legacy": legacy,
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceWorkbook": source_ref,
            "interactionRows": len(interactions),
            "physicsRows": len(physics),
        },
    }


def row_get(row, *keys):
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def build_glyphs(data_blocks, source_ref):
    map3 = data_blocks.get("Stability Engine v0.xlsx::MAP3", [])
    map2 = data_blocks.get("Stability Engine v0.xlsx::MAP2", [])
    rows = map3 + map2

    glyphs = []
    seen = set()

    for row in rows:
        node = row_get(row, "Node", "NodeName")
        if not node:
            continue
        node_str = clean_label(node)
        if not node_str:
            continue
        glyph_id = re.sub(r"[^a-z0-9]+", "-", node_str.lower()).strip("-")
        if not glyph_id or glyph_id in seen:
            continue
        seen.add(glyph_id)

        pressure = parse_float(row.get("ClusterPressure"))
        stability = parse_float(row.get("ClusterStabilityIndex"))

        if pressure is not None and pressure > 0.66:
            season = "storm"
        elif stability is not None and stability > 0.66:
            season = "soil"
        elif pressure is not None and pressure < 0.33:
            season = "fog"
        else:
            season = "shimmer"

        glyphs.append({
            "id": glyph_id,
            "node": node_str,
            "nodeType": clean_label(row.get("NodeType")),
            "chamber": clean_label(row.get("ClusterType") or row.get("Status") or "threshold") or "threshold",
            "season": season,
            "status": clean_label(row.get("Status")),
            "connectsTo": parse_edges(row.get("ConnectsTo")),
            "pressure": pressure,
            "stabilityIndex": stability,
        })

    return {
        "glyphs": glyphs,
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceWorkbook": source_ref,
            "glyphCount": len(glyphs),
        },
    }


def build_branches(data_blocks, source_ref):
    rows = []
    rows.extend(data_blocks.get("Stability Engine v0.xlsx::MAP3", []))
    rows.extend(data_blocks.get("Stability Engine v0.xlsx::MAP2", []))
    rows.extend(data_blocks.get("Stability Engine v0.xlsx::MAP_OLD", []))

    nodes = {}
    edges = set()

    for row in rows:
        node_raw = row_get(row, "Node", "NodeName")
        if not node_raw:
            continue
        node = clean_label(node_raw)
        if not node:
            continue

        node_id = re.sub(r"[^a-z0-9]+", "-", node.lower()).strip("-")
        status = clean_label(row.get("Status"))
        ntype = clean_label(row.get("NodeType"))

        pressure = parse_float(row.get("ClusterPressure"))
        stability = parse_float(row.get("ClusterStabilityIndex"))

        pressure_band = "low"
        if pressure is not None and pressure > 0.66:
            pressure_band = "high"
        elif pressure is not None and pressure >= 0.33:
            pressure_band = "mid"

        stability_band = "low"
        if stability is not None and stability > 0.66:
            stability_band = "high"
        elif stability is not None and stability >= 0.33:
            stability_band = "mid"

        nodes[node_id] = {
            "id": node_id,
            "nodeType": ntype,
            "status": status,
            "chamber": clean_label(row.get("ClusterType") or ntype or "threshold") or "threshold",
            "cluster": clean_label(row.get("Cluster") or row.get("AssignedCluster")),
            "pressureBand": pressure_band,
            "stabilityBand": stability_band,
        }

        for target in parse_edges(row.get("ConnectsTo")):
            target_id = re.sub(r"[^a-z0-9]+", "-", target.lower()).strip("-")
            if target_id:
                condition = status if status else "active"
                edges.add((node_id, target_id, condition))

    edge_rows = []
    for from_id, to_id, condition in sorted(edges):
        weight = 1.0
        if "rare" in condition.lower() or "locked" in condition.lower():
            weight = 0.6
        edge_rows.append({
            "from": from_id,
            "to": to_id,
            "condition": condition,
            "weight": weight,
        })

    return {
        "branches": {
            "nodes": sorted(nodes.values(), key=lambda x: x["id"]),
            "edges": edge_rows,
        },
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceWorkbook": source_ref,
            "nodeCount": len(nodes),
            "edgeCount": len(edge_rows),
        },
    }


def build_target_payload(target_name, data_blocks, source_ref):
    name = target_name.lower()
    if name.endswith("vault-dialogues.json"):
        return build_dialogues(data_blocks, source_ref)
    if name.endswith("vault-mythic.json"):
        return build_mythic(data_blocks, source_ref)
    if name.endswith("vault-glyphs.json"):
        return build_glyphs(data_blocks, source_ref)
    if name.endswith("vault-branches.json"):
        return build_branches(data_blocks, source_ref)

    return {
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceWorkbook": source_ref,
            "note": "No builder implemented for target",
        }
    }


def main():
    parser = argparse.ArgumentParser(description="Export runtime JSON from Excel workbooks using excel-export-map.json")
    parser.add_argument("--map", dest="map_path", default="scripts/excel-export-map.json", help="Path to export mapping JSON")
    parser.add_argument("--root", dest="root", default="", help="Override source workbook root directory")
    parser.add_argument("--out-root", dest="out_root", default=".", help="Website workspace root for target output paths")
    args = parser.parse_args()

    map_path = Path(args.map_path).resolve()
    if not map_path.exists():
        raise SystemExit("Map file not found: " + str(map_path))

    with map_path.open("r", encoding="utf-8") as f:
        export_map = json.load(f)

    root = Path(args.root).resolve() if args.root else Path(export_map.get("root", "")).resolve()
    if not root.exists():
        raise SystemExit("Excel source root not found: " + str(root))

    out_root = Path(args.out_root).resolve()
    source_priority = export_map.get("sourcePriority", [])
    targets = export_map.get("targets", [])

    workbook_index = build_workbook_index(root)
    wb_cache = {}

    written = []
    for target in targets:
        target_rel = target.get("target")
        if not target_rel:
            continue

        data_blocks = {}
        source_ref = []

        for source in target.get("sources", []):
            requested_wb = source.get("workbook", "")
            wb_path = resolve_workbook_path(requested_wb, workbook_index, source_priority, root)
            if not wb_path:
                continue

            wb = open_workbook(wb_path, wb_cache)
            sheet = source.get("sheet", "")
            columns = source.get("columns", [])
            rows = sheet_rows(wb, sheet, columns)

            block_key = f"{requested_wb}::{sheet}"
            data_blocks[block_key] = rows
            source_ref.append({"workbook": wb_path.name, "sheet": sheet, "rows": len(rows)})

        payload = build_target_payload(target_rel, data_blocks, source_ref)

        target_path = out_root / target_rel
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with target_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)

        written.append(str(target_path))

    for wb in wb_cache.values():
        wb.close()

    print("Export complete. Files written:")
    for item in written:
        print(" - " + item)


if __name__ == "__main__":
    main()
